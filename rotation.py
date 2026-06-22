import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import random
import re
import json
from datetime import date
from html import escape
from io import BytesIO
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# 1. 페이지 설정
st.set_page_config(page_title="GM Manager Central", layout="wide")

st.markdown("""
    <style>
    .meal-bg { background-color: #ffff00 !important; color: black !important; font-weight: bold; }
    </style>
    """, unsafe_allow_html=True)

st.title("GENTLE MONSTER 로테이션 시스템 v71.4")

# 🔗 매장 및 시트 설정
STORES = {
    "하우스 서울": {
        "ID": "19CvEiqbhPqNpz2KzcBQh7vVaH40O_ZuR6MFYdw98c5Q",
        "DAY_TYPES": {
            "평일": {"DB_GID": "738722894", "TO_GID": "410487706"},
            "주말": {"DB_GID": "0", "TO_GID": "2126973547"},
        },
    },
    "하우스 도산": {
        "ID": "1nqSbhCPnO1o_vRSubJCuLjbbZxmtRMjioTtA_ZzzNLc",
        "DAY_TYPES": {
            "평일": {"DB_GID": "0", "TO_GID": "2126973547"},
            "주말": {"DB_GID": "0", "TO_GID": "2126973547"},
        },
    },
    "테스트": {
        "ID": "1qZp0-8sqjLN65gbLPObPDYVzNDuWZiPlik6FJAM3MWY",
        "DAY_TYPES": {
            "평일": {"DB_GID": "0", "TO_GID": "2126973547"},
            "주말": {"DB_GID": "0", "TO_GID": "2126973547"},
        },
    },
}

selected_store = st.sidebar.selectbox("🏠 담당 매장 선택", list(STORES.keys()))
selected_day_type = st.sidebar.radio("📅 운영 구분", ["평일", "주말"], horizontal=True)

store_config = STORES[selected_store]
SHEET_ID = store_config["ID"]
day_type_config = store_config["DAY_TYPES"][selected_day_type]
DB_SHEET_GID = day_type_config["DB_GID"]
TO_SHEET_GID = day_type_config["TO_GID"]

@st.cache_data(ttl=1)
def load_sheet_data(sheet_id, gid):
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
    try:
        df = pd.read_csv(url, skip_blank_lines=True, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.fillna("").replace(r'\.0$', '', regex=True)
        return df
    except Exception as e:
        st.error(f"시트 로딩 실패: {e}")
        return pd.DataFrame()

db_df = load_sheet_data(SHEET_ID, DB_SHEET_GID)
to_df = load_sheet_data(SHEET_ID, TO_SHEET_GID)

if db_df.empty: st.stop()

def get_clean_time(val):
    val = str(val).strip()
    if not val: return None
    nums = re.findall(r'\d+', val)
    return f"{int(nums[0]):02d}:00" if nums else None

def get_hour_from_time(val, default=None):
    clean = get_clean_time(val)
    if not clean:
        return default
    return int(clean.split(":")[0])

def build_work_range(in_val, out_val, default_in=11, default_out=21):
    in_hr = get_hour_from_time(in_val, default_in)
    out_hr = get_hour_from_time(out_val, default_out)
    if in_hr is None or out_hr is None:
        return None, None, None
    if out_hr <= in_hr:
        return None, in_hr, out_hr
    return range(in_hr, out_hr), in_hr, out_hr

def is_counter_zone(zone_name):
    zone = str(zone_name).upper()
    return "카운터" in zone or "COUNTER" in zone or "1F-C" in zone or "2F-C" in zone

def is_flexible_zone(zone_name):
    zone = str(zone_name)
    return "유동" in zone

def get_zone_category(zone_name):
    zone = str(zone_name).upper()
    if is_counter_zone(zone_name):
        return "counter"
    if is_flexible_zone(zone_name):
        return "flex"
    if "2F" in zone or "2층" in zone:
        return "2f"
    if "1F" in zone or "1층" in zone:
        return "1f"
    return "other"

def get_floor_bucket(zone_name):
    if not zone_name:
        return None
    zone = str(zone_name)
    upper = zone.upper()
    if "1층" in zone or "1F" in upper:
        return "1f"
    if "2층" in zone or "2F" in upper:
        return "2f"
    if "카운터" in zone or "COUNTER" in upper:
        # default to whichever floor indicator is included
        if "2" in zone:
            return "2f"
        if "1" in zone:
            return "1f"
        return "counter"
    if "유동" in zone:
        if "2" in zone:
            return "2f"
        if "1" in zone:
            return "1f"
    return None

def get_zone_priority(zone_name):
    if is_counter_zone(zone_name):
        return 0
    if is_flexible_zone(zone_name):
        return 2
    return 1

def is_w_zone(zone_name):
    zone = str(zone_name).upper()
    return bool(re.search(r'(^|[^A-Z])W($|[^A-Z])', zone))

def is_photo_zone(zone_name):
    zone = str(zone_name).upper()
    return "PHOTO" in zone

def is_b1_zone(zone_name):
    zone = str(zone_name).upper()
    return "B1" in zone

def get_special_zone_group(zone_name):
    if is_counter_zone(zone_name):
        return "counter"
    if is_w_zone(zone_name):
        return "w"
    return None

def normalize_schedule_value(value):
    text = str(value)
    if text.strip() == "":
        return ""
    return text

def is_enabled_flag(value):
    return any(x in str(value).lower() for x in ['o', 'y', '1', 'v', '예'])

def pick_best_staff(zone_name, pool, previous_assignments):
    if not pool:
        return None

    candidates = list(pool)
    random.shuffle(candidates)

    if is_flexible_zone(zone_name):
        non_flexible_last = [
            name for name in candidates
            if not is_flexible_zone(previous_assignments.get(name))
        ]
        if non_flexible_last:
            candidates = non_flexible_last
    else:
        from_flexible_last = [
            name for name in candidates
            if is_flexible_zone(previous_assignments.get(name))
        ]
        if from_flexible_last:
            candidates = from_flexible_last

    non_consecutive = [
        name for name in candidates if previous_assignments.get(name) != zone_name
    ]
    if non_consecutive:
        candidates = non_consecutive

    return candidates[0] if candidates else None

# --- 기본 데이터 파싱 ---
def get_initial_staff(data):
    type_col = next((c for c in data.columns if '구분' in c), '구분')
    name_col = next((c for c in data.columns if '이름' in c), '이름')
    res = []
    for _, row in data.iterrows():
        name = str(row.get(name_col, "")).strip()
        stype = str(row.get(type_col, "")).strip()
        if name and any(kw in stype for kw in ['정직', '파트']):
            res.append({
                "original_name": name,
                "type": '정직' if '정직' in stype else '파트',
                "in": get_clean_time(row.get('출근시간', '11')),
                "out": get_clean_time(row.get('퇴근시간', '21')),
                "meal1": get_clean_time(row.get('점심', '')),
                "meal2": get_clean_time(row.get('저녁', '')),
                "meal_p": get_clean_time(row.get('식사시간', '')),
                "can_counter": is_enabled_flag(row.get('카운터여부', 'X')),
                "can_flexible": is_enabled_flag(row.get('유동여부', 'X')),
            })
    return res

raw_staff = get_initial_staff(db_df)

# --- 사이드바: 파트타이머 상세 조정 ---
st.sidebar.header("🕹️ 인원 관리")
pt_list = [s for s in raw_staff if s['type'] == '파트']

pt_input_defaults = {
    s["original_name"]: {
        "in": s["in"] or "11:00",
        "out": s["out"] or "21:00",
        "meal": s["meal_p"] or "12:00",
    }
    for s in pt_list
}
pt_input_signature = json.dumps(
    {
        "store": selected_store,
        "day_type": selected_day_type,
        "defaults": pt_input_defaults,
    },
    ensure_ascii=False,
    sort_keys=True,
)

if st.session_state.get("pt_input_signature") != pt_input_signature:
    for pt_name, defaults in pt_input_defaults.items():
        st.session_state[f"in_{pt_name}"] = defaults["in"]
        st.session_state[f"out_{pt_name}"] = defaults["out"]
        st.session_state[f"meal_{pt_name}"] = defaults["meal"]
    st.session_state["pt_input_signature"] = pt_input_signature

selected_pt_names = st.sidebar.multiselect("⏱️ 출근 파트타이머 선택", [s['original_name'] for s in pt_list], default=[s['original_name'] for s in pt_list])

final_staff_configs = []

# 1. 정직원 처리 (조 이름 포함)
for s in [x for x in raw_staff if x['type'] == '정직']:
    work_range, in_hr, out_hr = build_work_range(s['in'], s['out'])
    if work_range is None:
        continue
    tag = "(A조)" if in_hr <= 10 else "(B조)"
    s['display_name'] = f"{s['original_name']}{tag}"
    s['meals'] = list(set([m for m in [s['meal1'], s['meal2']] if m]))
    s['work_range'] = work_range
    s['in'] = f"{in_hr:02d}:00"
    s['out'] = f"{out_hr:02d}:00"
    final_staff_configs.append(s)

# 2. 파트타이머 처리 (조 이름 제외 + 사이드바 조정값 반영)
if selected_pt_names:
    st.sidebar.markdown("---")
    st.sidebar.subheader("📋 파트타이머 시간 조정")
    for pt_name in selected_pt_names:
        pt_origin = next(s for s in pt_list if s['original_name'] == pt_name)
        with st.sidebar.expander(f"👤 {pt_name}"):
            c1, c2 = st.columns(2)
            new_in = c1.text_input(f"출근", key=f"in_{pt_name}")
            new_out = c2.text_input(f"퇴근", key=f"out_{pt_name}")
            new_meal = st.text_input(f"식사", key=f"meal_{pt_name}")

            pt_copy = pt_origin.copy()
            work_range, in_hr, out_hr = build_work_range(new_in, new_out)

            if work_range is None:
                st.warning(f"{pt_name}: 출근/퇴근 시간을 다시 확인해 주세요.")
                continue

            pt_copy['display_name'] = pt_name # 조 태그 없음
            pt_copy['in'] = f"{in_hr:02d}:00"
            pt_copy['out'] = f"{out_hr:02d}:00"
            pt_copy['meal_p'] = get_clean_time(new_meal)
            pt_copy['meals'] = [pt_copy['meal_p']] if pt_copy['meal_p'] else []
            pt_copy['work_range'] = work_range
            final_staff_configs.append(pt_copy)

config_signature = json.dumps(
    {
        "store": selected_store,
        "day_type": selected_day_type,
        "staff": [
            {
                "name": s["display_name"],
                "type": s["type"],
                "in": s.get("in"),
                "out": s.get("out"),
                "meals": s.get("meals", []),
                "can_counter": s.get("can_counter", False),
                "can_flexible": s.get("can_flexible", False),
            }
            for s in final_staff_configs
        ],
    },
    ensure_ascii=False,
    sort_keys=True,
)

if st.session_state.get("config_signature") != config_signature:
    st.session_state.pop("result_df", None)
    st.session_state["config_signature"] = config_signature

# --- 로테이션 엔진 ---
def run_rotation():
    working_names = [s['display_name'] for s in final_staff_configs]
    all_time_slots = [f"{h:02d}:00" for h in range(11, 21)]
    schedule_df = pd.DataFrame(index=all_time_slots, columns=working_names).fillna("-")
    all_zones = [c for c in to_df.columns if c != to_df.columns[0]]

    # TO 시트에 실제 존재하는 유동 구역만 사용
    flex_zones_in_to = [z for z in all_zones if is_flexible_zone(z)]
    flex_1f = next((z for z in flex_zones_in_to if "1" in z), None)
    flex_2f = next((z for z in flex_zones_in_to if "2" in z), None)

    staff_lookup = {s['display_name']: s for s in final_staff_configs}
    previous_assignments = {n: None for n in working_names}
    floor_state = {n: {"floor": None, "count": 0} for n in working_names}
    floor_1f_total = {n: 0 for n in working_names}  # 1층 총 배정 횟수
    special_zone_history = {n: {} for n in working_names}
    w_zone_hours = {n: 0 for n in working_names}
    counter_assignment_total = {n: 0 for n in working_names}
    counter_consecutive_hours = {n: 0 for n in working_names}
    MAX_1F = 3
    MAX_W_HOURS = 2
    MAX_CONSECUTIVE_COUNTER_HOURS = 2

    def get_zone_identity(zone_name):
        return str(zone_name).strip().upper()

    def is_part_timer(name):
        staff = staff_lookup.get(name, {})
        return staff.get("type") == "파트"

    def is_full_time(name):
        staff = staff_lookup.get(name, {})
        return staff.get("type") == "정직"

    def can_assign_zone(name, zone_name):
        special_group = get_special_zone_group(zone_name)
        zone_identity = get_zone_identity(zone_name)

        if special_group:
            previous_zone = special_zone_history[name].get(special_group)
            if previous_zone and previous_zone != zone_identity:
                return False

        if is_w_zone(zone_name) and w_zone_hours[name] >= MAX_W_HOURS:
            return False

        if is_counter_zone(zone_name) and counter_consecutive_hours[name] >= MAX_CONSECUTIVE_COUNTER_HOURS:
            return False

        return True

    def record_zone_assignment(name, zone_name):
        special_group = get_special_zone_group(zone_name)
        if special_group:
            special_zone_history[name][special_group] = get_zone_identity(zone_name)
        if is_w_zone(zone_name):
            w_zone_hours[name] += 1
        if is_counter_zone(zone_name):
            counter_assignment_total[name] += 1
            counter_consecutive_hours[name] += 1
        else:
            counter_consecutive_hours[name] = 0

    def pick_counter_staff(zone_name, candidates):
        if not candidates:
            return None

        prioritized = [n for n in candidates if is_full_time(n)]
        working_candidates = prioritized or candidates
        min_counter_total = min(counter_assignment_total[n] for n in working_candidates)
        least_used_candidates = [
            n for n in working_candidates
            if counter_assignment_total[n] == min_counter_total
        ]
        min_counter_streak = min(counter_consecutive_hours[n] for n in least_used_candidates)
        streak_balanced_candidates = [
            n for n in least_used_candidates
            if counter_consecutive_hours[n] == min_counter_streak
        ]
        return pick_best_staff(
            zone_name,
            streak_balanced_candidates,
            previous_assignments,
        )

    def update_floor_state(name, zone):
        floor = get_floor_bucket(zone)
        state = floor_state[name]
        if not floor:
            state["floor"] = None
            state["count"] = 0
            return
        if floor == "1f":
            floor_1f_total[name] += 1
        if state["floor"] == floor:
            state["count"] += 1
        else:
            state["floor"] = floor
            state["count"] = 1

    def can_assign_same_floor(name, floor):
        if not floor:
            return True
        state = floor_state[name]
        return not (state["floor"] == floor and state["count"] >= 2)

    for slot in all_time_slots:
        hr = int(slot.split(":")[0])
        pool = []
        for s in final_staff_configs:
            if slot in s["meals"]:
                schedule_df.at[slot, s['display_name']] = "식사"
                floor_state[s['display_name']]['floor'] = None
                floor_state[s['display_name']]['count'] = 0
                counter_consecutive_hours[s['display_name']] = 0
            elif hr in s["work_range"]:
                pool.append(s['display_name'])
            else:
                schedule_df.at[slot, s['display_name']] = " "
                floor_state[s['display_name']]['floor'] = None
                floor_state[s['display_name']]['count'] = 0
                counter_consecutive_hours[s['display_name']] = 0
        
        random.shuffle(pool)
        to_row = to_df[to_df[to_df.columns[0]].str.contains(slot, na=False)]
        
        if not to_row.empty:
            active_zones = [z for z in all_zones if str(to_row[z].iloc[0]).strip() != "0"]
            sorted_active_zones = sorted(active_zones, key=lambda z: (get_zone_priority(z), all_zones.index(z)))

            photo_zones = [z for z in sorted_active_zones if is_photo_zone(z)]
            prioritized_zones = [z for z in sorted_active_zones if z not in photo_zones]
            sorted_active_zones = photo_zones + prioritized_zones

            for z in sorted_active_zones:
                raw = str(to_row[z].iloc[0]).strip()
                mi = int(raw.split('-')[0]) if '-' in raw else int(float(raw or 0))
                assigned = 0
                while assigned < mi:
                    zone_is_1f = get_floor_bucket(z) == "1f"
                    eligible = [
                        n for n in pool
                        if not (is_counter_zone(z) and not staff_lookup[n]["can_counter"])
                        and not (is_flexible_zone(z) and not staff_lookup[n]["can_flexible"])
                        and not (zone_is_1f and floor_1f_total[n] >= MAX_1F)
                        and can_assign_zone(n, z)
                    ]
                    if is_b1_zone(z):
                        part_timer_eligible = [n for n in eligible if is_part_timer(n)]
                        if part_timer_eligible:
                            eligible = part_timer_eligible
                    zone_floor = get_floor_bucket(z)
                    floor_filtered = [n for n in eligible if can_assign_same_floor(n, zone_floor)]
                    working_candidates = floor_filtered or eligible
                    if is_counter_zone(z):
                        chosen = pick_counter_staff(z, working_candidates)
                    else:
                        chosen = pick_best_staff(
                            z,
                            working_candidates,
                            previous_assignments,
                        )
                    if not chosen and is_photo_zone(z) and assigned == 0:
                        photo_fallback = [
                            n for n in pool
                            if can_assign_zone(n, z)
                        ]
                        chosen = pick_best_staff(
                            z,
                            photo_fallback,
                            previous_assignments,
                        )
                    if not chosen:
                        break

                    schedule_df.at[slot, chosen] = z
                    previous_assignments[chosen] = z
                    record_zone_assignment(chosen, z)
                    assigned += 1
                    pool.remove(chosen)
                    update_floor_state(chosen, z)

            flexible_pool = [n for n in pool if staff_lookup[n]["can_flexible"]]
            inflexible_pool = [n for n in pool if not staff_lookup[n]["can_flexible"]]

            for n in flexible_pool: # TO 시트에 있는 유동 구역만 사용
                if not flex_1f and not flex_2f:
                    # 유동 구역 자체가 없으면 미배정
                    schedule_df.at[slot, n] = "-"
                    previous_assignments[n] = None
                    floor_state[n]["floor"] = None
                    floor_state[n]["count"] = 0
                    continue

                current_assignments = [
                    val for val in schedule_df.loc[slot].tolist()
                    if str(val).strip() not in ["-", "", " ", "식사"]
                ]
                f1_cnt = sum(1 for val in current_assignments if get_zone_category(val) == "1f")
                f2_cnt = sum(1 for val in current_assignments if get_zone_category(val) == "2f")

                can_1f = flex_1f and floor_1f_total[n] < MAX_1F
                if can_1f and flex_2f:
                    flexible_zone = flex_1f if f1_cnt <= f2_cnt else flex_2f
                    if previous_assignments.get(n) == flexible_zone:
                        flexible_zone = flex_2f if flexible_zone == flex_1f else flex_1f
                    # 1층 선택됐는데 한도 초과면 2층으로
                    if flexible_zone == flex_1f and floor_1f_total[n] >= MAX_1F:
                        flexible_zone = flex_2f
                elif flex_2f:
                    flexible_zone = flex_2f
                elif can_1f:
                    flexible_zone = flex_1f
                else:
                    flexible_zone = flex_2f  # 한도 초과 시 2층으로 강제

                schedule_df.at[slot, n] = flexible_zone
                previous_assignments[n] = flexible_zone
                record_zone_assignment(n, flexible_zone)
                update_floor_state(n, flexible_zone)

            for n in inflexible_pool:
                schedule_df.at[slot, n] = "-"
                previous_assignments[n] = None
                floor_state[n]["floor"] = None
                floor_state[n]["count"] = 0
                counter_consecutive_hours[n] = 0
    return schedule_df

if st.sidebar.button("🚀 로테이션 자동 생성", width="stretch"):
    st.session_state.result_df = run_rotation()

# --- 화면 출력 ---
if 'result_df' in st.session_state:
    res = st.session_state.result_df
    st.write(f"### 📅 [{selected_store} / {selected_day_type}] 로테이션")
    st.caption("수정은 아래 표에서 하고, 변경 내용은 모바일 공유용 현황판에 바로 반영됩니다.")
    display_df = res.transpose().map(normalize_schedule_value)
    display_df.index.name = "직원명"
    editor_df = display_df.reset_index()
    editor_df = editor_df[["직원명"] + [c for c in editor_df.columns if c != "직원명"]]
    zone_columns = [c for c in to_df.columns if c != to_df.columns[0]]
    zone_choices = set(zone_columns)
    zone_choices.update(str(val).strip() for val in display_df.values.flatten() if str(val).strip())
    zone_choices.update(["식사", "1층 유동", "2층 유동", "-", ""])
    zone_choices = sorted(zone_choices)
    column_settings = {
        col: (
            st.column_config.SelectboxColumn(options=zone_choices)
            if col != "직원명"
            else st.column_config.TextColumn(label="직원명", disabled=True)
        )
        for col in editor_df.columns
    }
    edited_editor_df = st.data_editor(
        editor_df,
        width="stretch",
        height=450,
        column_config=column_settings,
        hide_index=True,
        num_rows="fixed",
        key="rotation_editor",
    )
    edited_df = edited_editor_df.copy()
    edited_df["직원명"] = edited_df["직원명"].astype(str).str.strip()
    edited_df = edited_df.set_index("직원명")
    edited_df.index.name = "직원명"
    edited_df = edited_df.reindex(columns=display_df.columns)
    edited_df = edited_df.map(normalize_schedule_value)
    csv_bytes = edited_df.to_csv(index=True).encode('utf-8')
    file_name = f"rotation_{selected_store}_{selected_day_type}_{date.today():%Y%m%d}"

    def parse_required_count(raw_value):
        raw = str(raw_value).strip()
        if raw in ["", "0", "-", "nan"]:
            return 0
        return int(raw.split('-')[0]) if '-' in raw else int(float(raw or 0))

    def build_unassigned_zone_summary(df):
        summary_rows = []
        total_empty_zones = 0
        total_shortfall = 0
        affected_times = 0

        for slot in df.columns:
            to_row = to_df[to_df[to_df.columns[0]].astype(str).str.contains(slot, na=False)]
            if to_row.empty:
                continue

            required_counts = {}
            for zone in zone_columns:
                required = parse_required_count(to_row[zone].iloc[0])
                if required > 0:
                    required_counts[zone] = required

            assigned_counts = {}
            for value in df[slot].tolist():
                zone_name = str(value).strip()
                if zone_name in required_counts:
                    assigned_counts[zone_name] = assigned_counts.get(zone_name, 0) + 1

            empty_zones = []
            shortage_zones = []
            shortfall_count = 0
            for zone, required in required_counts.items():
                assigned = assigned_counts.get(zone, 0)
                if required > 0 and assigned == 0:
                    empty_zones.append(zone)
                shortfall = max(required - assigned, 0)
                if shortfall > 0:
                    shortage_zones.append(f"{zone} ({shortfall})")
                    shortfall_count += shortfall

            empty_count = len(empty_zones)
            if empty_count > 0 or shortfall_count > 0:
                affected_times += 1
            total_empty_zones += empty_count
            total_shortfall += shortfall_count

            summary_rows.append(
                {
                    "time": slot,
                    "empty_count": empty_count,
                    "shortfall_count": shortfall_count,
                    "status": "주의" if empty_count > 0 or shortfall_count > 0 else "정상",
                    "empty_zones": empty_zones,
                    "shortage_zones": shortage_zones,
                }
            )

        return summary_rows, total_empty_zones, total_shortfall, affected_times

    def get_staff_color(name):
        s_info = next((s for s in final_staff_configs if s['display_name'] == name), None)
        if not s_info:
            return "#111827"
        if "(A조)" in name:
            return "#f97316"
        if "(B조)" in name:
            return "#2563eb"
        if s_info["type"] == '정직':
            return "#1d4ed8"
        return "#059669"

    def get_zone_background(value):
        text = str(value)
        low = text.lower()
        if "카운터" in low or "counter" in low:
            return "#ede9fe"
        if "2층" in low or "2f" in low:
            return "#fee2e2"
        if "1층" in low or "1f" in low:
            return "#dbeafe"
        return ""

    def excel_color(hex_color):
        return hex_color.replace("#", "").upper()

    def style_rotation_worksheet(ws, df):
        thin_side = Side(style="thin", color="DDDDDD")
        header_fill = PatternFill(fill_type="solid", fgColor=excel_color("#f8f9fa"))
        meal_fill = PatternFill(fill_type="solid", fgColor=excel_color("#fff5ba"))
        center_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = center_alignment
            cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for row_idx, staff in enumerate(df.index, start=2):
            name_cell = ws.cell(row=row_idx, column=1)
            name_cell.font = Font(bold=True, color=excel_color(get_staff_color(staff)))
            name_cell.alignment = center_alignment
            name_cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

            for col_idx, value in enumerate(df.loc[staff], start=2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = center_alignment
                cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

                if str(value) == "식사":
                    cell.fill = meal_fill
                    continue

                zone_color = get_zone_background(value)
                if zone_color:
                    cell.fill = PatternFill(fill_type="solid", fgColor=excel_color(zone_color))

        ws.freeze_panes = "B2"
        ws.column_dimensions["A"].width = 18

        for col_idx in range(2, len(df.columns) + 2):
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[column_letter].width = 12

    with BytesIO() as buf:
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            edited_df.to_excel(writer, index=True, sheet_name="rotation")
            style_rotation_worksheet(writer.book["rotation"], edited_df)
        buf.seek(0)
        excel_bytes = buf.getvalue()

    def build_table(df):
        table_html = "<div class='table-scroll'><table class='rotation-table'>"
        table_html += "<thead><tr><th>직원</th>"
        for time in df.columns:
            table_html += f"<th>{escape(str(time))}</th>"
        table_html += "</tr></thead><tbody>"
        for staff, row in df.iterrows():
            color = get_staff_color(staff)
            table_html += f"<tr><td class='staff-name' style='color: {color};'>{escape(str(staff))}</td>"
            for _, val in row.items():
                text = normalize_schedule_value(val)
                bg = ""
                if text == "식사":
                    bg = "background-color: #fff5ba;"
                else:
                    zone_color = get_zone_background(text)
                    if zone_color:
                        bg = f"background-color: {zone_color};"
                table_html += f"<td style='{bg}'>{escape(text)}</td>"
            table_html += "</tr>"
        table_html += "</tbody></table></div>"
        return table_html

    def build_unassigned_zone_panel(summary_rows):
        panel_html = "<div class='gap-board'>"
        for row in summary_rows:
            status_class = "alert" if row["empty_count"] > 0 or row["shortfall_count"] > 0 else "ok"
            empty_detail = ", ".join(row["empty_zones"]) if row["empty_zones"] else "없음"
            shortage_detail = ", ".join(row["shortage_zones"]) if row["shortage_zones"] else "없음"
            panel_html += (
                f"<div class='gap-card {status_class}'>"
                f"<div class='gap-time'>{escape(row['time'])}</div>"
                f"<div class='gap-status'>{escape(row['status'])}</div>"
                f"<div class='gap-count'>빈 구역 {row['empty_count']}개</div>"
                f"<div class='gap-count'>부족 인원 {row['shortfall_count']}명</div>"
                f"<div class='gap-detail'>빈 구역: {escape(empty_detail)}</div>"
                f"<div class='gap-detail'>부족 현황: {escape(shortage_detail)}</div>"
                "</div>"
            )
        panel_html += "</div>"
        return panel_html

    table_styles = (
        "<style>"
        ".table-scroll{overflow:auto;background:#fff;border:1px solid #ddd;border-radius:12px;}"
        ".rotation-table{width:max-content;min-width:100%;border-collapse:collapse;font-size:0.95rem;}"
        ".rotation-table th,.rotation-table td{border:1px solid #ddd;padding:8px;text-align:center;vertical-align:middle;white-space:nowrap;}"
        ".rotation-table thead th{position:sticky;top:0;background:#f8f9fa;z-index:3;}"
        ".rotation-table .staff-name{position:sticky;left:0;background:#fff;font-weight:700;z-index:2;}"
        ".rotation-table thead th:first-child{left:0;z-index:4;}"
        ".gap-board{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px;margin:12px 0 20px;}"
        ".gap-card{border-radius:14px;padding:14px 16px;border:1px solid #d1d5db;background:#fff;box-shadow:0 6px 18px rgba(15,23,42,0.06);}"
        ".gap-card.ok{background:linear-gradient(180deg,#f0fdf4 0%,#ffffff 100%);border-color:#bbf7d0;}"
        ".gap-card.alert{background:linear-gradient(180deg,#fff7ed 0%,#ffffff 100%);border-color:#fdba74;}"
        ".gap-time{font-size:0.95rem;font-weight:700;color:#111827;}"
        ".gap-status{margin-top:6px;font-size:0.85rem;font-weight:700;}"
        ".gap-card.ok .gap-status{color:#15803d;}"
        ".gap-card.alert .gap-status{color:#c2410c;}"
        ".gap-count{margin-top:6px;font-size:0.92rem;color:#111827;}"
        ".gap-detail{margin-top:8px;font-size:0.84rem;line-height:1.45;color:#4b5563;}"
        "</style>"
    )
    gap_summary_rows, total_empty_zones, total_shortfall, affected_times = build_unassigned_zone_summary(edited_df)
    table_html = build_table(edited_df)
    gap_panel_html = build_unassigned_zone_panel(gap_summary_rows)
    page_html = "<!doctype html><html lang='ko'><head><meta charset='utf-8'/><title>모바일 공유 현황판</title>"
    page_html += (
        "<style>"
        "html,body{height:100%;margin:0;padding:0;background:#f8fafc;font-family:'Pretendard','Noto Sans KR',sans-serif;}"
        ".page-wrap{display:flex;flex-direction:column;height:100%;padding:16px;box-sizing:border-box;gap:12px;}"
        "h1{margin:0;font-size:1.4rem;}"
        "</style>"
        f"{table_styles}"
    )
    page_html += "</head><body><div class='page-wrap'><h1>모바일 공유 현황판</h1>"
    page_html += "<h2 style='margin:0;font-size:1.05rem;'>미배정 구역 체크</h2>"
    page_html += gap_panel_html
    page_html += table_html
    page_html += "</div></body></html>"

    widget_html = f"""
    <div style='margin-bottom:8px;'>
        <button style='border:0; padding:10px 16px; font-weight:600; background:#111827; color:#fff; border-radius:8px; cursor:pointer;'
                onclick='openLargeRotation()'>🖥️ 크게 보기</button>
    </div>
    <script>
    const largeRotationContent = {json.dumps(page_html)};
    function openLargeRotation() {{
        const win = window.open('', '_blank');
        if (!win) return;
        win.document.write(largeRotationContent);
        win.document.close();
    }}
    </script>
    """
    st.markdown(table_styles, unsafe_allow_html=True)
    st.markdown("### 🚨 미배정 구역 체크")
    metric_col1, metric_col2, metric_col3 = st.columns(3)
    metric_col1.metric("빈 구역 총수", total_empty_zones)
    metric_col2.metric("부족 인원 총수", total_shortfall)
    metric_col3.metric("영향 시간대", affected_times)
    st.markdown(gap_panel_html, unsafe_allow_html=True)
    st.write("---")
    st.markdown("### 📸 모바일 공유용 현황판")
    components.html(widget_html, height=110)
    st.write("---")
    st.markdown("### 📥 다운로드")
    st.download_button("📥 현재 배정 다운로드 (CSV)", data=csv_bytes, file_name=f"{file_name}.csv", mime="text/csv")
    st.download_button(
        "📥 현재 배정 다운로드 (Excel)",
        data=excel_bytes,
        file_name=f"{file_name}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
